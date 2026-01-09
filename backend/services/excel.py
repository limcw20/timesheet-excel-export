from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta
import os
from pathlib import Path

# Environment Variables
env_path = os.getenv("FILE_DIRECTORY")
timesheet_filename = os.getenv("TIMESHEET_FILENAME")
staff_name = os.getenv("STAFF_NAME")
staff_ref_no = os.getenv("STAFF_REF_NO")


# Save file to directory from home path
target_path = Path.home() / env_path

target_path.mkdir(parents=True, exist_ok=True)

def generate_timesheet_excel(timesheet: dict) -> str:

    wb = Workbook()
    ws = wb.active

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bottom_border = Border(
        bottom=Side(style='thin')
    )

    # Column widths
    ws.column_dimensions["A"].width = 11
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 13
    ws.column_dimensions["H"].width = 9
    ws.column_dimensions["I"].width = 11
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 18

    # 1st Row Title
    ws.merge_cells("A1:K1") 
    title_cell = ws["A1"] 
    title_cell.value = "Time Sheet" 
    title_cell.font = Font(bold=True, size=20) 
    title_cell.alignment = center_align

    # Date Range Calculation
    start_year = timesheet['start_year']
    start_month = timesheet['start_month']
    start_date = datetime(start_year, start_month, 4) # start from nth day of month

    if start_month == 12:
        next_year = start_year + 1
        next_month = 1
    else:
        next_year = start_year
        next_month = start_month + 1

    end_date = datetime(next_year, next_month, 3) # end at nth day of next month
    total_days = (end_date - start_date).days + 1

    entries = {e.get('day_of_month'): e for e in timesheet.get('entries', [])}
    
    # 3rd Row - employee ref, date period, name of employee
    ws.merge_cells("A3:C3")
    employee_ref_cell = ws["A3"]
    employee_ref_cell.value = f"Contract Reference No: {staff_ref_no}"
    employee_ref_cell.font = Font(bold=True, size=12)

    ws.merge_cells("D3:F3")
    date_period_cell = ws["D3"]
    date_period_cell.value = f"Date Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"
    date_period_cell.font = Font(bold=True, size=12)
    date_period_cell.alignment = center_align
    
    ws.merge_cells("G3:K3")
    name_cell = ws["G3"]
    name_cell.value = f"Name of Staff: {staff_name}"
    name_cell.font = Font(bold=True, size=12)
    name_cell.alignment = Alignment(horizontal="right")
    
    # Headers
    headers = ["Day", "Date", "Reason for Absence", "Time In", "Time Out", "Hours Worked","OT From","OT To","OT hours","Total Hours", "OT Approved By"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header) # Header at row n
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    # Row Highlights
    weekend_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid") # Grey
    holiday_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid") # Light Orangey kind of colour
    leave_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Light Blue

    for i in range(total_days):
        row_num = i + 7 # Data starts from row n
        date_obj = start_date + timedelta(days=i)
        weekday = date_obj.strftime("%A")
        day_of_month = date_obj.day

        # Default values for each iteration
        hours_worked = 8
        reason_for_absence = None
        time_in = "09:00am"
        time_out = "06:00pm"
        ot_from = "-"
        ot_to = "-"
        ot_hours = "-"
        ot_approved_by = "-"
        total_hours = hours_worked

        # Extract payload data if exists
        if day_of_month in entries:
            entry = entries[day_of_month]
            hours_worked = entry.get('hours_worked', hours_worked)
            time_in = entry.get('time_in', time_in)
            time_out = entry.get('time_out', time_out)
            reason_for_absence = entry.get('reason_for_absence', reason_for_absence)
            ot_from = entry.get('ot_from', ot_from)
            ot_to = entry.get('ot_to', ot_to)
            ot_hours = entry.get('ot_hours', ot_hours)
            ot_approved_by = entry.get('ot_approved_by', ot_approved_by)
            total_hours = int(hours_worked + ot_hours)
            

        # Default logic if weekends
        if weekday in ["Saturday", "Sunday"]:
            hours_worked = None
            reason_for_absence = None
            time_in = None
            time_out = None
            ot_from = None
            ot_to = None
            ot_hours = None
            total_hours = None
            ot_approved_by = None
            for col in range(1, 12):
                ws.cell(row=row_num, column=col).fill = weekend_fill
                
        # Logic for full day absent codes
        elif reason_for_absence in ["AL", "MC", "UPL","PH"]:
            hours_worked = None
            reason_for_absence = reason_for_absence
            time_in = None
            time_out = None
            ot_from = None
            ot_to = None
            ot_hours = None
            total_hours = None
            ot_approved_by = None
            if reason_for_absence == "PH":
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).fill = holiday_fill
            else:
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).fill = leave_fill
            
        
        # Logic for half day absent codes
        elif reason_for_absence in ["AM leave", "PM leave", "half day"]:
            hours_worked = 4
            reason_for_absence = reason_for_absence
            ot_from = "-"
            ot_to = "-"
            ot_hours = "-"
            total_hours = hours_worked
            ot_approved_by = "-"
            if reason_for_absence == "AM leave":
                time_in = "01:00pm"
                time_out = "05:00pm"
            else:
                time_in = "09:00am"
                time_out = "01:00pm"
            
            if reason_for_absence == "half day":
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).fill = holiday_fill
            else:
                for col in range(1, 12):
                    ws.cell(row=row_num, column=col).fill = leave_fill

        # Fill in rows on specified columns with data(value)
        ws.cell(row=row_num, column=1, value=weekday)
        ws.cell(row=row_num, column=2, value=date_obj.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=3, value=reason_for_absence)
        ws.cell(row=row_num, column=4, value=time_in)
        ws.cell(row=row_num, column=5, value=time_out)
        ws.cell(row=row_num, column=6, value=hours_worked)
        ws.cell(row=row_num, column=7, value=ot_from)
        ws.cell(row=row_num, column=8, value=ot_to)
        ws.cell(row=row_num, column=9, value=ot_hours)
        ws.cell(row=row_num, column=10, value=total_hours)
        ws.cell(row=row_num, column=11, value=ot_approved_by)

        for col in range(1, 12):
            ws.cell(row=row_num, column=col).alignment = center_align
            ws.cell(row=row_num, column=col).border = thin_border
            
        
    # Name & Date of preparer at the bottom
    preparer_cell = ws["B" + str(row_num + 2)]
    preparer_cell.value = "Name:"
    preparer_cell.alignment = Alignment(horizontal="right")
    preparer_fill_cell = ws["C" + str(row_num + 2)]
    preparer_fill_cell.value = staff_name
    preparer_fill_cell.border = bottom_border
    preparer_fill_cell.alignment = center_align
    
    preparer_date_cell = ws["B" + str(row_num + 5)]
    preparer_date_cell.value = "Date:"
    preparer_date_cell.alignment = Alignment(horizontal="right")
    preparer_date_fill_cell = ws["C" + str(row_num + 5)]
    preparer_date_fill_cell.value = datetime.now().strftime("%Y-%m-%d")
    preparer_date_fill_cell.alignment = center_align
    preparer_date_fill_cell.border = bottom_border
    
    # Name, Date, Signature of Approver at the bottom
    approver_cell = ws["E" + str(row_num + 2)]
    approver_cell.value = "Supervisor Name:"
    approver_cell.alignment = Alignment(horizontal="right")
    approver_fill_cell_f = ws["F" + str(row_num + 2)]
    approver_fill_cell_g = ws["G" + str(row_num + 2)]
    approver_fill_cell_f.border = bottom_border
    approver_fill_cell_g.border = bottom_border
    
    approver_signature_cell = ws["E" + str(row_num + 5)]
    approver_signature_cell.value = "Signature:"
    approver_signature_cell.alignment = Alignment(horizontal="right")
    approver_signature_fill_cell = ws["F" + str(row_num + 5)]
    approver_signature_fill_cell.border = bottom_border
    
    approver_date_cell = ws["E" + str(row_num + 7)]
    approver_date_cell.value = "Date:"
    approver_date_cell.alignment = Alignment(horizontal="right")
    approver_date_fill_cell = ws["F" + str(row_num + 7)]
    approver_date_fill_cell.border = bottom_border
    
    out_file = os.path.join(str(target_path), f"{timesheet_filename}.xlsx")
    wb.save(out_file)
    return out_file
